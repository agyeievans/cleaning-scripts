import csv
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# File paths
input_file = 'Amharic News Dataset 2.csv'  # Replace with your input CSV file path
output_csv = 'cleaned_output.csv'  # Replace with your desired output CSV file path
output_excel = 'Amharic News Dataset 2.xlsx'  # Replace with your desired Excel file path

# Step 1: Clean the CSV file
with open(input_file, 'r', encoding='utf-8-sig') as infile, open(output_csv, 'w', newline='', encoding='utf-8-sig') as outfile:
    reader = csv.reader(infile)
    writer = csv.writer(outfile)
    
    for row in reader:
        cleaned_row = [cell.strip() for cell in row]
        writer.writerow(cleaned_row)

# Step 2: Convert cleaned CSV to Excel
df = pd.read_csv(output_csv)
df.to_excel(output_excel, index=False, engine='openpyxl')

# Step 3: Set default row heights and column widths
workbook = load_workbook(output_excel)
sheet = workbook.active

# Set column widths (default 15 for all columns)
default_column_width = 15
for col in range(1, sheet.max_column + 1):
    sheet.column_dimensions[get_column_letter(col)].width = default_column_width

# Set row heights (default 20 for all rows)
default_row_height = 20
for row in range(1, sheet.max_row + 1):
    sheet.row_dimensions[row].height = default_row_height

# Save the formatted Excel file
workbook.save(output_excel)

print(f"Cleaned and formatted Excel file has been saved as {output_excel}")
